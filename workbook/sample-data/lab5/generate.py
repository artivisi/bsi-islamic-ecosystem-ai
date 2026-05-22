"""
Generator data fiktif untuk Lab 5 — Workshop Planning & Portfolio.

Membuat 3 file RKAP XLSX yang sengaja inkonsisten antar dept (unit, periode,
nama kolom, struktur sheet) supaya peserta belajar normalisasi dan flagging
sebelum konsolidasi.

Catatan kunci:
- Semua nama lembaga = lembaga publik yang sudah dikenal luas (no real BSI
  prospect data).
- Angka fiktif; tidak mencerminkan target / realisasi BSI sesungguhnya.

Jalankan: python3 generate.py
Output: rkap-dept-{1,2,3}.xlsx, realisasi-q1-2027.csv, agenda-radir.md
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent

HEADER_FILL = PatternFill(start_color="2E3192", end_color="2E3192", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, max_width=50):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines():
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), max_width)


# ============================================================================
# DEPT 1 — Pendidikan & Ulama
# Segmen: Pesantren, Sekolah Islam, PT Islam
# Format: 1 sheet, kolom Bahasa Indonesia, unit Rp Miliar, phasing kuartalan
# ============================================================================

def build_dept1():
    wb = Workbook()
    ws = wb.active
    ws.title = "RKAP 2027"

    headers = [
        "Segmen",
        "Target Pembiayaan 2027 (Rp Miliar)",
        "Target DPK 2027 (Rp Miliar)",
        "Target Fee Based 2027 (Rp Miliar)",
        "Q1 (%)", "Q2 (%)", "Q3 (%)", "Q4 (%)",
        "Jumlah Prospek Target",
        "Capex 2027 (Rp Juta)",
        "Opex 2027 (Rp Juta)",
        "Asumsi Kunci",
    ]
    ws.append(headers)

    rows = [
        ["Pesantren",       850, 620, 45, 20, 25, 30, 25, 120,  850, 3200,
         "Akuisisi 30 pesantren baru via approach induk ormas (NU, Muhammadiyah, Persis). Asumsi closing rate 25% dari pipeline 120 prospek."],
        ["Sekolah Islam",   420, 380, 28, 15, 30, 35, 20,  95,  480, 2100,
         "Pertumbuhan dari yayasan multi-unit (SD-SMA-SMK satu yayasan). Fokus paket cash management + payroll guru."],
        ["PT Islam",        380, 290, 32, 25, 30, 25, 20,  45,  620, 1850,
         "Fokus 10 PT prioritas (UIN, IAIN, universitas berbasis ormas). Paket BSI Mobile bagi mahasiswa jadi entry product."],
    ]
    for row in rows:
        ws.append(row)

    style_header(ws)
    autosize(ws)
    ws.freeze_panes = "B2"

    wb.save(OUT / "rkap-dept-1.xlsx")


# ============================================================================
# DEPT 2 — Layanan & Sosial
# Segmen: RS Islam, Masjid, ZISWAF
# Format: 3 sheet (Target / Budget / Asumsi), kolom campur Eng-Ind,
#         unit Rp Juta (BEDA dengan Dept 1!), tahunan saja (no phasing)
# ============================================================================

def build_dept2():
    wb = Workbook()

    ws = wb.active
    ws.title = "Target"
    ws.append([
        "Segment",
        "Total Target 2027 (Rp Juta)",
        "Lending Target (Rp Juta)",
        "Deposit Target (Rp Juta)",
        "FBI (Rp Juta)",
    ])
    target_rows = [
        ["RS Islam", 1850000, 1100000,  680000, 70000],
        ["Masjid",     95000,    5000,   85000,  5000],
        ["ZISWAF",    220000,   12000,  195000, 13000],
    ]
    for row in target_rows:
        ws.append(row)
    style_header(ws)
    autosize(ws)

    ws2 = wb.create_sheet("Budget")
    ws2.append([
        "Segment",
        "Capex (Rp Juta)",
        "Opex (Rp Juta)",
    ])
    budget_rows = [
        ["RS Islam", 1200, 4500],
        ["Masjid",    180,  950],
        ["ZISWAF",    320, 1450],
    ]
    for row in budget_rows:
        ws2.append(row)
    style_header(ws2)
    autosize(ws2)

    ws3 = wb.create_sheet("Asumsi")
    ws3.append(["Segment", "Pipeline Prospect", "Conversion Asumsi (%)", "Notes"])
    asumsi_rows = [
        ["RS Islam",  35, 40, "Ekspansi ke 12 RS Islam baru, fokus cash management + pembiayaan infrastructure rumah sakit. Anchor: RS Islam Sultan Agung Semarang, RSI Jakarta."],
        ["Masjid",   180,  8, "Segmen low-funding, fokus mass payment + ZISWAF digital. Conversion rate rendah karena bukan revenue-driver utama."],
        ["ZISWAF",    60, 25, "Target overlap dengan Masjid — perlu dikaji ulang. Beberapa nazhir dan amil terdaftar di kategori Masjid juga ZISWAF."],
    ]
    for row in asumsi_rows:
        ws3.append(row)
    style_header(ws3)
    autosize(ws3)

    wb.save(OUT / "rkap-dept-2.xlsx")


# ============================================================================
# DEPT 3 — Ekonomi & Komunitas
# Segmen: Halal F&B/Fashion, Travel Umrah, Ormas Islam
# Format: 1 sheet, kolom English snake_case, unit IDR ribuan (BEDA LAGI!),
#         phasing semester (H1 / H2)
# ============================================================================

def build_dept3():
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan 2027"

    headers = [
        "segment",
        "financing_target_idr_k",
        "third_party_funds_target_idr_k",
        "fee_target_idr_k",
        "h1_pct",
        "h2_pct",
        "prospect_count",
        "capex_idr_k",
        "opex_idr_k",
        "key_assumption",
    ]
    ws.append(headers)

    # Unit: ribuan rupiah. 320_000_000 ribu = 320 miliar.
    rows = [
        ["Halal F&B/Fashion", 320_000_000, 180_000_000, 25_000_000, 45, 55,  80,  420_000, 1_850_000,
         "Sertifikasi halal MUI jadi entry point; fokus UMKM yang scaling ke retail modern. Asumsi: regulasi wajib halal 2027 mendorong demand pembiayaan modal kerja."],
        ["Travel Umrah",      280_000_000, 510_000_000, 18_000_000, 60, 40,  65,  280_000, 1_650_000,
         "Q3-Q4 peak musim haji & umrah. Integrasi BSI Hajj. Funding mix: deposit dominant (jamaah savings) sebelum keberangkatan."],
        ["Ormas Islam",        95_000_000, 145_000_000,  8_000_000, 50, 50,  25,  180_000,   920_000,
         "Approach lewat MoU level induk ormas (NU, Muhammadiyah, Persis, Mathla'ul Anwar). Konversi dari MoU ke transaksi historically 18 bulan."],
    ]
    for row in rows:
        ws.append(row)

    style_header(ws)
    autosize(ws)

    wb.save(OUT / "rkap-dept-3.xlsx")


# ============================================================================
# REALISASI Q1 2027 — untuk Sesi 5.4
# CSV ternormalisasi (sudah dalam Rp Miliar) — instruktur asumsikan ada tim
# operasional yang sudah pre-process. Variasi tier dan catatan early warning.
# ============================================================================

def build_realisasi():
    headers = [
        "segmen",
        "dept_pengelola",
        "target_pembiayaan_q1_miliar",
        "realisasi_pembiayaan_q1_miliar",
        "achievement_pct_pembiayaan",
        "target_dpk_q1_miliar",
        "realisasi_dpk_q1_miliar",
        "achievement_pct_dpk",
        "target_fee_q1_miliar",
        "realisasi_fee_q1_miliar",
        "achievement_pct_fee",
        "jumlah_prospek_aktif_q1",
        "perubahan_prospek_vs_q4_2026",
        "catatan_q1",
    ]
    rows = [
        ["Pesantren",        "Dept 1 — Pendidikan", 170.0, 155.0,  91, 124.0, 118.0, 95, 9.0, 8.5, 94,  118,  -2,
         "Hijau overall, tapi laju akuisisi prospek baru melambat di Maret (Jan 45 → Feb 38 → Mar 35). Mulai velocity drop."],
        ["Sekolah Islam",    "Dept 1 — Pendidikan",  63.0,  48.0,  76,  57.0,  44.0, 77, 4.2, 3.1, 74,   67, -28,
         "Kuning. Pipeline prospek aktif turun 30% dari Q4 2026 (95 → 67). Yayasan multi-unit yang dijanjikan jadi anchor belum closing."],
        ["PT Islam",         "Dept 1 — Pendidikan",  95.0,  62.0,  65,  72.5,  55.0, 76, 8.0, 4.2, 53,   42,  -3,
         "Merah pembiayaan. 1 deal besar dengan UIN Jakarta (estimasi 25 miliar) pending closing → concentration risk. Jika gagal, achievement turun ke ~38%."],
        ["RS Islam",         "Dept 2 — Layanan",    275.0, 305.0, 111, 170.0, 144.5, 85, 17.5, 16.8, 96,  38,   3,
         "Hijau pembiayaan tapi DPK underperform (85%) → funding mix imbalance. Ekspansi RS Islam terlalu cepat sebelum landasan deposit settled."],
        ["Masjid",           "Dept 2 — Layanan",     1.25,   0.9,  72,  21.25, 16.0, 75, 1.25, 0.95, 76, 152, -18,
         "Kuning. Masjid Istiqlal yang diharapkan jadi anchor cash management belum tanda-tangan MoU. Pipeline 152 masjid aktif tapi conversion lambat."],
        ["ZISWAF",           "Dept 2 — Layanan",     3.0,   3.7, 123,  48.75, 67.0, 137, 3.25, 4.1, 126,  58,   8,
         "Hijau, surge karena program Ramadan ZISWAF digital. WARNING: perlu antisipasi pasca-Ramadan slowdown di Q2."],
        ["Halal F&B/Fashion","Dept 3 — Ekonomi",     36.0,  28.5,  79,  20.25, 14.5, 72, 2.8, 2.0, 71,   62, -18,
         "Kuning. Sertifikasi halal MUI delay 3 bulan, pipeline UMKM tertunda. Tidak fatal — momentum akan kembali Q2 jika sertifikasi resume."],
        ["Travel Umrah",     "Dept 3 — Ekonomi",     42.0,  35.0,  83,  76.5,  72.0, 94, 2.7, 2.5, 93,   58,   5,
         "Kuning di pembiayaan, tapi Q1 memang seasonal low (musim umrah baru aktif Q3). DPK kuat karena jamaah savings stabil. Bukan red flag."],
        ["Ormas Islam",      "Dept 3 — Ekonomi",      11.875, 7.2,  61,  18.125, 11.5, 63, 1.0, 0.6, 60,  17,  -8,
         "Merah. Belum ada MoU induk ormas signed (NU, Muhammadiyah masih negosiasi). Pipeline 25 prospek tidak konversi tanpa payung MoU dulu."],
    ]

    with open(OUT / "realisasi-q1-2027.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


# ============================================================================
# AGENDA RADIR — fiktif, bulan berjalan
# ============================================================================

def build_agenda_radir():
    content = """# Agenda Radir Bulanan ISE BD — Mei 2027

*Forum:* Rapat Direksi BSI
*Tanggal:* Senin, 24 Mei 2027, 09:00 – 11:30
*Tempat:* Ruang Direksi, BSI Tower Lantai 21
*Peserta:* Direktur Utama, Direktur Wholesale, Direktur Retail, SEVP Islamic Ecosystem
*Pemilik agenda:* Kepala Dept Planning & Portfolio ISE BD

---

## 1. Review Pencapaian Q1 2027 Portfolio ISE BD

- Status traffic light per segmen (9 segmen, 3 dept BD pengelola)
- Highlight segmen merah: PT Islam, Ormas Islam
- Highlight surge: ZISWAF (Ramadan effect) — perlu antisipasi pasca-Ramadan
- Slide pendukung: tracker realisasi Q1, chart distribusi achievement

## 2. Persetujuan Revisi Target Q2 untuk Segmen Merah

- *PT Islam:* concentration risk pada deal UIN Jakarta — opsi A: tetap target, opsi B: realokasi 15 miliar ke Pesantren. Rekomendasi Planning: opsi B.
- *Ormas Islam:* MoU induk ormas pending — opsi A: tunda target Q2, opsi B: shift fokus ke ormas tingkat wilayah dulu. Rekomendasi Planning: opsi B.
- *Decision needed:* persetujuan revisi target atau pertahankan.

## 3. Update Progress MoU dengan Induk Ormas

- NU: pertemuan tingkat PBNU dijadwalkan minggu ke-3 Juni
- Muhammadiyah: dialog masih di level Majelis Ekonomi PP
- Persis & Mathla'ul Anwar: menunggu inisiasi
- Implikasi cross-segment: MoU induk membuka akses ke Pesantren, Sekolah, PT, dan Ormas sekaligus

## 4. Persetujuan Tambahan Capex untuk Ekspansi RS Islam Q2

- RS Islam achievement 111% pembiayaan tapi DPK 85% → butuh tim relationship manager tambahan untuk garap deposit side
- Permintaan: tambahan capex 800 juta untuk hire 4 RM khusus segmen RS Islam
- Justifikasi: jika tidak, funding mix imbalance akan compound di Q2-Q3
- *Decision needed:* approve / tunda / partial approval

## 5. Strategi Mitigasi Keterlambatan Sertifikasi Halal MUI

- Halal F&B/Fashion segmen achievement 79% — akar masalah: sertifikasi halal MUI delay 3 bulan
- Pipeline UMKM tertahan menunggu sertifikat
- Rekomendasi Planning: ajukan dialog formal BSI dengan BPJPH untuk percepatan
- *Decision needed:* surat resmi dari Direksi ke BPJPH (Q2)

## 6. Update Proyek BSI Mobile untuk Segmen PT Islam

- 10 PT prioritas, status integrasi BSI Mobile dengan SIM kampus
- Sudah live di 4 PT, 3 dalam UAT, 3 masih perencanaan
- Issue: PT Islam belum reach scale yang diharapkan — perlu pivot ke fokus payroll dulu sebelum mahasiswa
- *Decision needed:* lanjut roadmap atau pivot

## 7. Inisiatif Cross-Segment: Cash Management untuk Yayasan Multi-Unit

- Insight lintas segmen: banyak yayasan punya pesantren + sekolah + RS sekaligus (mis. Yayasan Sukmana, Yayasan Pondok Modern Gontor, Yayasan Hasyim Asy'ari)
- Saat ini 3 dept BD pegang relationship terpisah dengan unit-unit yayasan sama
- Usulan: paket cash management lintas-unit dengan single point of contact
- *Decision needed:* pilot dengan 3 yayasan anchor di Q3

## 8. Lain-lain & Penutup

- Komunikasi internal: ringkasan keputusan dikirim ke 3 dept BD dalam 2 hari kerja
- Radir berikutnya: Senin 28 Juni 2027
"""
    (OUT / "agenda-radir.md").write_text(content, encoding="utf-8")


if __name__ == "__main__":
    build_dept1()
    build_dept2()
    build_dept3()
    build_realisasi()
    build_agenda_radir()
    print("Generated:")
    for path in sorted(OUT.glob("*")):
        if path.name != "generate.py" and path.name != "README.md":
            print(f"  {path.name}")
